import openpyxl
import os

def crear_vehiculo():
    codigo = input("Ingrese el código del vehículo: ")
    marca = input("Ingrese la marca del vehículo: ")
    modelo = input("Ingrese el modelo del vehículo: ")
    precio = float(input("Ingrese el precio del vehículo: "))
    kilometraje = int(input("Ingrese el kilometraje del vehículo: "))

    return codigo, marca, modelo, precio, kilometraje

def agregar_vehiculo(archivo, vehiculo):
    hoja = archivo["listado"]
    hoja.append(vehiculo)
    archivo.save("vehiculos.xlsx")
    print("Vehículo agregado exitosamente.")

def editar_vehiculo(archivo, codigo):
    hoja = archivo["listado"]

    for fila in hoja.iter_rows(min_row=2, max_col=1, max_row=hoja.max_row):
        if fila[0].value == codigo:
            print("Vehículo encontrado. Proporcione la nueva información:")
            nuevo_vehiculo = crear_vehiculo()
            for i, valor in enumerate(nuevo_vehiculo, start=2):
                fila[i - 1].value = valor
            archivo.save("vehiculos.xlsx")
            print("Vehículo editado exitosamente.")
            return

    print("Vehículo no encontrado.")

def eliminar_vehiculo(archivo, codigo):
    hoja = archivo["listado"]

    for fila in hoja.iter_rows(min_row=2, max_col=1, max_row=hoja.max_row):
        if fila[0].value == codigo:
            hoja.delete_rows(fila[0].row)
            archivo.save("vehiculos.xlsx")
            print("Vehículo eliminado exitosamente.")
            return

    print("Vehículo no encontrado.")

def listar_vehiculos(archivo):
    hoja = archivo["listado"]

    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row):
        for celda in fila:
            print(celda.value, end="\t")
        print()

def carga_masiva(archivo, ruta_archivo_masivo):
    try:
        with open(ruta_archivo_masivo, "r") as file:
            lineas = file.readlines()
            for linea in lineas:
                datos = linea.strip().split("|")
                agregar_vehiculo(archivo, datos)
        print("Carga masiva completada.")
    except FileNotFoundError:
        print("Archivo masivo no encontrado.")

def main():
    if not os.path.exists("vehiculos.xlsx"):
        # Crear el archivo de Excel por si este no existe xd 
        archivo = openpyxl.Workbook()
        hoja = archivo.active
        hoja.title = "listado"
        hoja.append(["Codigo", "Marca", "Modelo", "Precio", "Kilometraje"])
        archivo.save("vehiculos.xlsx")
    
    archivo = openpyxl.load_workbook("vehiculos.xlsx")

    while True:
        print("\n*** Menú de Mantenimiento de Vehículos ***")
        print("1. Crear Vehículo")
        print("2. Editar Vehículo")
        print("3. Eliminar Vehículo")
        print("4. Listar Vehículos")
        print("5. Carga Masiva de Vehículos")
        print("0. Salir")

        opcion = input("Ingrese la opción deseada: ")

        if opcion == "1":
            nuevo_vehiculo = crear_vehiculo()
            agregar_vehiculo(archivo, nuevo_vehiculo)
        elif opcion == "2":
            codigo = input("Ingrese el código del vehículo a editar: ")
            editar_vehiculo(archivo, codigo)
        elif opcion == "3":
            codigo = input("Ingrese el código del vehículo a eliminar: ")
            eliminar_vehiculo(archivo, codigo)
        elif opcion == "4":
            listar_vehiculos(archivo)
        elif opcion == "5":
            ruta_archivo_masivo = input("Ingrese la ruta del archivo masivo: ")
            carga_masiva(archivo, ruta_archivo_masivo)
        elif opcion == "0":
            break
        else:
            print("Opción no válida. Inténtelo de nuevo.")

    archivo.close()

if __name__ == "__main__":
    main()
